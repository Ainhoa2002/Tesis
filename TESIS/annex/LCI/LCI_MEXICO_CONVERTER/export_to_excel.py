#!/usr/bin/env python3
"""Export subsystem results to Excel workbooks.

This script reuses subsystem selection from add_eliminate_component.py.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import sys
import logging

# Helper to preserve interactive prints when running in a TTY,
# but emit structured logging when running non-interactively.
_IS_TTY = sys.stdout.isatty()
def _out(msg: str, level: str = "info") -> None:
    if _IS_TTY:
        print(msg)
    else:
        getattr(logging, level)(msg)

from tools.add_eliminate_component import (
    MAX_SELECTION_ATTEMPTS,
    SelectionAborted,
    choose_subsystem,
    discover_subsystem_files,
)

BASE_DIR = Path(__file__).parent
DEFAULT_EXPORT_DIR = BASE_DIR.parent


def load_csv_optional(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    if not path.exists():
        return [], []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    return headers, rows


def load_consolidated_mass_results(base_dir: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    headers: List[str] = []
    rows: List[Dict[str, str]] = []

    for csv_path in sorted(base_dir.glob("*_component_results.csv")):
        subsystem = csv_path.name[: -len("_component_results.csv")]
        file_headers, file_rows = load_csv_optional(csv_path)
        if not file_headers:
            continue

        merged_headers = list(file_headers)
        if "Subsystem" not in merged_headers:
            merged_headers.append("Subsystem")

        for header in merged_headers:
            if header not in headers:
                headers.append(header)

        for row in file_rows:
            new_row = dict(row)
            new_row["Subsystem"] = subsystem
            rows.append(new_row)

    return headers, rows


def discover_section_files(base_dir: Path) -> Dict[str, Path]:
    """Discover all SECTION_*_ipe_flows_from_parameters.csv files."""
    sections: Dict[str, Path] = {}
    for csv_path in sorted(base_dir.glob("SECTION_*_ipe_flows_from_parameters.csv")):
        # Extract section name from filename: SECTION_<name>_ipe_flows_from_parameters.csv
        section_name = csv_path.name[len("SECTION_"): -len("_ipe_flows_from_parameters.csv")]
        sections[section_name] = csv_path
    return sections


def write_sheet(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def prompt_output_directory(default_dir: Path) -> Path:
    while True:
        folder_input = input(
            f"\nOutput folder (Enter for default: {default_dir}): "
        ).strip()

        if not folder_input:
            return default_dir

        chosen_dir = Path(folder_input).expanduser()
        if not chosen_dir.is_absolute():
            chosen_dir = (Path.cwd() / chosen_dir).resolve()

        if chosen_dir.exists() and chosen_dir.is_dir():
            return chosen_dir

        _out(f"Folder not found: {chosen_dir}", level="warning")


def prompt_output_filename(default_name: str) -> str:
    while True:
        file_input = input(
            f"Output filename (Enter for default: {default_name}): "
        ).strip()

        if not file_input:
            return default_name

        if not file_input.lower().endswith(".xlsx"):
            file_input = file_input + ".xlsx"

        return file_input


def choose_export_mode() -> str:
    _out("\nExport mode:")
    _out("  1. Export one subsystem")
    _out("  2. Export all subsystems")
    _out("  3. Export sections")

    attempts = 0
    while True:
        raw = input("Mode [1/2/3]: ").strip().lower()
        if raw in {"1", "one", "single", "subsystem"}:
            return "one"
        if raw in {"2", "all", "todo", "todos", "*"}:
            return "all"
        if raw in {"3", "sections", "section"}:
            return "sections"

        attempts += 1
        _out("Invalid option. Enter 1, 2, or 3.", level="warning")
        if attempts >= MAX_SELECTION_ATTEMPTS:
            raise SelectionAborted("Too many invalid attempts. Operation canceled.")


def export_all_subsystems_to_excel(
    base_dir: Path,
    subsystems: Dict[str, Path],
    output_dir: Path,
) -> Tuple[List[Path], List[str]]:
    exported_paths: List[Path] = []
    skipped_subsystems: List[str] = []
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    _out("\nChoose output filename for each subsystem (Enter for default).")
    for subsystem_name in sorted(subsystems.keys()):
        default_filename = f"{subsystem_name}_results_export_{stamp}.xlsx"
        _out(f"\nSubsystem: {subsystem_name}")
        output_filename = prompt_output_filename(default_filename)

        output = export_subsystem_results_to_excel(
            base_dir,
            subsystem_name,
            output_dir,
            output_filename,
        )

        if output is None:
            skipped_subsystems.append(subsystem_name)
            continue

        exported_paths.append(output)

    return exported_paths, skipped_subsystems


def export_subsystem_results_to_excel(
    base_dir: Path,
    subsystem: str,
    output_dir: Path,
    output_filename: str,
) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        _out("openpyxl is required for Excel export.", level="error")
        _out("Install it with: .\\.venv\\Scripts\\python.exe -m pip install openpyxl", level="error")
        return None

    sources = [
        ("Parameters", base_dir / f"{subsystem}_component_parameters.csv"),
        ("Mass_Results", base_dir / f"{subsystem}_component_mass_results.csv"),
        ("Component_IO", base_dir / f"{subsystem}_component_io_flows.csv"),
        ("Grouped_Flows", base_dir / f"{subsystem}_ipe_flows_from_parameters.csv"),
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    exported = 0

    for sheet_name, csv_path in sources:
        headers, rows = load_csv_optional(csv_path)
        if not headers:
            continue
        ws = workbook.create_sheet(sheet_name[:31])
        write_sheet(ws, headers, rows)
        exported += 1

    if exported == 0:
        _out("No data found to export for this subsystem.", level="warning")
        return None

    output_path = output_dir / output_filename
    workbook.save(output_path)
    return output_path


def export_total_bom_to_excel(
    base_dir: Path,
    output_dir: Path,
    output_filename: str,
) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        _out("openpyxl is required for Excel export.", level="error")
        _out("Install it with: .\\.venv\\Scripts\\python.exe -m pip install openpyxl", level="error")
        return None

    workbook = Workbook()
    workbook.remove(workbook.active)
    exported = 0

    sources = [
        ("Parameters_All", base_dir / "component_library_parameters_all.csv"),
        ("Mass_Results_All", None),
        ("Ecoinvent_Totals", base_dir / "component_library_ecoinvent_totals.csv"),
    ]

    for sheet_name, csv_path in sources:
        if sheet_name == "Mass_Results_All":
            headers, rows = load_consolidated_mass_results(base_dir)
        else:
            headers, rows = load_csv_optional(csv_path) if csv_path is not None else ([], [])
        if not headers:
            continue
        ws = workbook.create_sheet(sheet_name[:31])
        write_sheet(ws, headers, rows)
        exported += 1

    if exported == 0:
        _out("No data found to export for total BoM.", level="warning")
        return None

    output_path = output_dir / output_filename
    workbook.save(output_path)
    return output_path


def export_sections_to_excel(
    base_dir: Path,
    output_dir: Path,
    output_filename: str,
) -> Path | None:
    """Export all sections to a single Excel workbook, one section per sheet."""
    try:
        from openpyxl import Workbook
    except ImportError:
        _out("openpyxl is required for Excel export.", level="error")
        _out("Install it with: .\\.venv\\Scripts\\python.exe -m pip install openpyxl", level="error")
        return None

    sections = discover_section_files(base_dir)
    
    if not sections:
        _out("No section files found to export.", level="warning")
        return None

    workbook = Workbook()
    workbook.remove(workbook.active)
    exported = 0

    for section_name in sorted(sections.keys()):
        csv_path = sections[section_name]
        headers, rows = load_csv_optional(csv_path)
        
        if not headers:
            continue
        
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = section_name[:31]
        ws = workbook.create_sheet(sheet_name)
        write_sheet(ws, headers, rows)
        exported += 1

    if exported == 0:
        _out("No data found to export for sections.", level="warning")
        return None

    output_path = output_dir / output_filename
    workbook.save(output_path)
    return output_path


def write_export_readme(
    output_dir: Path,
    exported_items: List[Tuple[Path, str]],
) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    readme_path = output_dir / f"export_readme_{stamp}.txt"
    lines = [
        "# Resumen de export",
        "",
        "Archivos exportados:",
    ]

    for file_path, description in exported_items:
        lines.append(f"- {file_path.name}: {description}")

    readme_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return readme_path


def main() -> None:
    try:
        subsystems = discover_subsystem_files(BASE_DIR)
        export_mode = choose_export_mode()

        if export_mode == "one":
            subsystem_name, _ = choose_subsystem(subsystems)
            export_dir = prompt_output_directory(DEFAULT_EXPORT_DIR)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"{subsystem_name}_results_export_{stamp}.xlsx"
            export_filename = prompt_output_filename(default_filename)

            output = export_subsystem_results_to_excel(
                BASE_DIR,
                subsystem_name,
                export_dir,
                export_filename,
            )
            if output is not None:
                _out(f"\nExport completed: {output}")
                readme_path = write_export_readme(
                    export_dir,
                    [
                        (
                            output,
                            "Excel de subsistema con hojas: Parameters, Mass_Results, Component_IO, Grouped_Flows.",
                        )
                    ],
                )
                _out(f"Readme creado: {readme_path}")
            return

        if export_mode == "sections":
            export_dir = prompt_output_directory(DEFAULT_EXPORT_DIR)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"sections_export_{stamp}.xlsx"
            export_filename = prompt_output_filename(default_filename)

            output = export_sections_to_excel(
                BASE_DIR,
                export_dir,
                export_filename,
            )
            if output is not None:
                _out(f"\nSections export completed: {output}")
                readme_path = write_export_readme(
                    export_dir,
                    [
                        (
                            output,
                            "Excel con hojas por sección, cada hoja contiene los flows de la sección con columnas: Flow, UUID, Unit, Amount, Direction, Section, Subsections, Source_subsystems, Component_rows, Total_mass_kg, Missing_process_components, UUID_provider.",
                        )
                    ],
                )
                _out(f"Readme creado: {readme_path}")
            return

        export_dir = prompt_output_directory(DEFAULT_EXPORT_DIR)
        exported_paths, skipped_subsystems = export_all_subsystems_to_excel(
            BASE_DIR,
            subsystems,
            export_dir,
        )

        _out("\nBoM total export file (includes parameters_all, mass_results_all and ecoinvent_totals):")
        bom_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_bom_filename = f"bom_total_export_{bom_stamp}.xlsx"
        bom_filename = prompt_output_filename(default_bom_filename)
        bom_output = export_total_bom_to_excel(
            BASE_DIR,
            export_dir,
            bom_filename,
        )

        if exported_paths:
            _out("\nExport completed for:")
            for path in exported_paths:
                _out(f"  - {path}")
        else:
            _out("\nNo subsystem exports were generated.")

        if bom_output is not None:
            _out("\nBoM total export completed:")
            _out(f"  - {bom_output}")
        else:
            _out("\nBoM total export was not generated.")

        if skipped_subsystems:
            _out("\nSubsystems with no exportable data:")
            for subsystem_name in skipped_subsystems:
                _out(f"  - {subsystem_name}")

        readme_items: List[Tuple[Path, str]] = []
        for subsystem_output in exported_paths:
            readme_items.append(
                (
                    subsystem_output,
                    "Excel de subsistema con hojas: Parameters, Mass_Results, Component_IO, Grouped_Flows.",
                )
            )

        if bom_output is not None:
            readme_items.append(
                (
                    bom_output,
                    "Excel de BoM total con hojas: Parameters_All, Mass_Results_All y Ecoinvent_Totals.",
                )
            )

        if readme_items:
            readme_path = write_export_readme(export_dir, readme_items)
            _out(f"\nReadme creado: {readme_path}")
    except SelectionAborted as exc:
        _out(str(exc), level="error")


if __name__ == "__main__":
    main()
