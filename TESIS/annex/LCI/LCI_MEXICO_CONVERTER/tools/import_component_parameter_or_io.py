"""
Role: Import component parameter files or I/O definitions into the converter.

Brief: Utility to read and normalize component parameter CSVs and I/O tables
so they can be consumed by the converter pipeline.
"""

#!/usr/bin/env python3
"""
Import component parameters from Excel into CSV format or duplicate an existing CSV.
Excel must have the same columns as the output CSV (Designators, Section, Subsection, etc.).
This allows building component databases directly in Excel and importing to the pipeline.
"""

from __future__ import annotations

import csv
import shutil
from pathlib import Path

import openpyxl
import sys
import logging

# interactive-safe output helper
_IS_TTY = sys.stdout.isatty()
# Purpose: Out.
def _out(msg: str, level: str = "info") -> None:
    if _IS_TTY:
        print(msg)
    else:
        getattr(logging, level)(msg)


# Purpose: Ensure csv suffix.
def _ensure_csv_suffix(name: str) -> str:
    if not name.lower().endswith(".csv"):
        return f"{name}.csv"
    return name


# Purpose: Is yes answer.
def _is_yes_answer(answer):
    return answer.strip().casefold() in {"y", "yes", "s", "si", "sí"}


# Purpose: Prompt import mode.
def _prompt_import_mode():
    _out("\nImport options:")
    _out("1. Excel (BoM parameters)")
    _out("2. Duplicate .csv")

    while True:
        option = input("Choose option [1/2]: ").strip().lower()
        if option in {"1", "excel", "x"}:
            return "excel"
        if option in {"2", "csv", "duplicate", "dup"}:
            return "csv"
        _out("Invalid option. Enter 1 or 2.", level="warning")


# Purpose: Prompt directory.
def _prompt_directory(prompt_text, default_dir):
    while True:
        dir_input = input(prompt_text).strip()
        chosen_dir = Path(dir_input) if dir_input else default_dir
        if chosen_dir.exists() and chosen_dir.is_dir():
            return chosen_dir
        _out(f"Folder not found: {chosen_dir}", level="warning")


# Purpose: Prompt output csv path.
def _prompt_output_csv_path(default_dir, default_name, disallow_path=None):
    while True:
        csv_prompt = f"\nOutput CSV filename (Enter for default: {default_name}): "
        csv_name = input(csv_prompt).strip()
        if not csv_name:
            csv_name = default_name

        if not csv_name:
            _out("Filename cannot be empty", level="warning")
            continue

        csv_name = _ensure_csv_suffix(csv_name)

        output_csv = default_dir / csv_name
        if disallow_path is not None and output_csv.resolve() == disallow_path.resolve():
            _out("Output CSV must be different from source CSV.", level="warning")
            continue

        if output_csv.exists():
            overwrite_answer = input(
                f"File already exists: {output_csv}. Overwrite? [y/N]: "
            )
            if not _is_yes_answer(overwrite_answer):
                _out("Choose a different filename.", level="info")
                continue
        return output_csv


# Purpose: Resolve source csv.
def _resolve_source_csv(source_dir, source_name, extra_search_dirs=None):
    source_path = Path(_ensure_csv_suffix(source_name))
    if source_path.is_absolute():
        return source_path if source_path.exists() and source_path.is_file() else None

    primary_candidate = source_dir / source_path
    if primary_candidate.exists() and primary_candidate.is_file():
        return primary_candidate

    for search_dir in extra_search_dirs or []:
        fallback_candidate = search_dir / source_path
        if fallback_candidate.exists() and fallback_candidate.is_file():
            return fallback_candidate

    return None


# Purpose: Choose csv from directory.
def _choose_csv_from_directory(source_dir):
    csv_files = sorted(p for p in source_dir.glob("*.csv") if p.is_file())
    if not csv_files:
        return None

    _out("\nCSV files in folder:")
    for i, path in enumerate(csv_files, start=1):
        _out(f"  {i}. {path.name}")

    while True:
        raw = input("Choose CSV number (Enter to type filename manually): ").strip()
        if not raw:
            return None
        try:
            idx = int(raw)
        except ValueError:
            _out("Invalid number. Enter one of the listed options.", level="warning")
            continue

        if 1 <= idx <= len(csv_files):
            return csv_files[idx - 1]

        _out("Invalid number. Enter one of the listed options.", level="warning")


# Purpose: Prompt source csv path.
def _prompt_source_csv_path(default_dir, extra_search_dirs=None):
    while True:
        dir_prompt = f"Source CSV folder (Enter for default: {default_dir}): "
        source_dir = _prompt_directory(dir_prompt, default_dir)

        selected_csv = _choose_csv_from_directory(source_dir)
        if selected_csv is not None:
            return selected_csv

        source_name = input("Source CSV filename: ").strip()
        if not source_name:
            _out("Filename cannot be empty", level="warning")
            continue

        source_csv = _resolve_source_csv(source_dir, source_name, extra_search_dirs)
        if source_csv is not None:
            if source_csv.parent.resolve() != source_dir.resolve():
                _out(f"Source CSV found in: {source_csv.parent}")
            return source_csv

        filename = source_name if source_name.lower().endswith(".csv") else f"{source_name}.csv"
        _out(f"CSV file not found in: {source_dir / filename}", level="warning")
        if extra_search_dirs:
            _out("Also checked fallback folders.")


# Purpose: Duplicate csv.
def duplicate_csv(source_csv, output_csv):
    shutil.copyfile(source_csv, output_csv)


# Purpose: Select sheet name.
def _select_sheet_name(workbook, requested_name=None):
    if requested_name:
        if requested_name in workbook.sheetnames:
            return requested_name
        raise KeyError(requested_name)

    if "Parameters" in workbook.sheetnames:
        return "Parameters"
    if "Sheet1" in workbook.sheetnames:
        return "Sheet1"
    if len(workbook.sheetnames) == 1:
        return workbook.sheetnames[0]
    return workbook.sheetnames[0]


# Purpose: Import from excel.
def import_from_excel(workbook_path, output_csv, sheet_name=None):
    """
    Import component parameters from Excel to CSV.
    -Path to Excel file with component data
    -Output CSV path
    -sheet_name: Name of the sheet to read (default: auto-detect)
    
    Returns:
        Tuple with imported row count and the sheet name used.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    try:
        resolved_sheet_name = _select_sheet_name(wb, sheet_name)
        ws = wb[resolved_sheet_name]
    except KeyError:
        requested_label = sheet_name if sheet_name else "<auto-detect>"
        _out(f"Error: Sheet '{requested_label}' could not be resolved in workbook", level="error")
        _out(f"Available sheets: {wb.sheetnames}", level="error")
        return 0, None
    
    header_row = []
    for cell_value in ws.iter_rows(min_row=1, max_row=1, values_only=True).__next__():
        if cell_value is None:
            break
        header_row.append(str(cell_value).strip())
    
    if not header_row:
        _out("Error: Excel file has no headers in first row", level="error")
        return 0, None

    _out(f"Found {len(header_row)} columns: {', '.join(header_row[:5])}...")
    
    rows = []

    for r in range(2, ws.max_row + 1):
        row_data = {}
        has_content = False

        for col_idx, field_name in enumerate(header_row, start=1):
            cell_value = ws.cell(r, col_idx).value
            if cell_value is not None:
                row_data[field_name] = str(cell_value).strip()
                has_content = True
            else:
                row_data[field_name] = ""

        if has_content:
            rows.append(row_data)
    
    # Write CSV with same headers as read from Excel
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=header_row)
        writer.writeheader()
        writer.writerows(rows)
    
    return len(rows), resolved_sheet_name


# Purpose: Main.
def main():
    mass_calc_dir = Path(__file__).parent.resolve()
    current_git_dir = Path.cwd().resolve()
    default_input_dir = mass_calc_dir
    
    _out("IMPORT COMPONENT PARAMETERS")
    mode = _prompt_import_mode()

    if mode == "excel":
        # Ask for base directory where Excel workbook is located.
        input_dir = _prompt_directory(
            f"Folder path (Enter for default: {default_input_dir}): ",
            default_input_dir,
        )

        # Ask for workbook name (with or without .xlsx extension).
        while True:
            workbook_name = input("\n Excel file name (e.g., BoM_fuse_card or BoM_fuse_card.xlsx): ").strip()
            if not workbook_name:
                _out("File name cannot be empty", level="warning")
                continue

            if not workbook_name.lower().endswith(".xlsx"):
                workbook_name = f"{workbook_name}.xlsx"

            excel_path = input_dir / workbook_name
            if excel_path.exists():
                break
            _out(f"Excel file not found: {excel_path}", level="warning")

        workbook_stem = excel_path.stem
        default_csv_name = (
            f"{workbook_stem[4:]}_component_parameters.csv"
            if workbook_stem.lower().startswith("bom_")
            else f"{workbook_stem}_component_parameters.csv"
        )

        output_csv = _prompt_output_csv_path(mass_calc_dir, default_csv_name)

        # Ask for sheet name
        sheet_name = input("\nSheet name (Enter for auto-detect): ").strip()
        if not sheet_name:
            sheet_name = None

        _out("")
        _out(f"Importing from: {excel_path}")
        _out(f"Sheet: {sheet_name or 'auto-detect'}")
        _out(f"Output: {output_csv}")
        _out("")

        count, resolved_sheet_name = import_from_excel(excel_path, output_csv, sheet_name)

        if resolved_sheet_name is None:
            _out("\nNo CSV file was generated.")
            return

        _out(f"\nImported {count} component rows")
        _out(f"Sheet used: {resolved_sheet_name}")
        _out(f"Output saved to: {output_csv}")
        return

    source_csv = _prompt_source_csv_path(
        mass_calc_dir,
        extra_search_dirs=[current_git_dir],
    )
    if source_csv is None:
        _out("No source CSV selected. Aborting.", level="warning")
        return
    output_dir = _prompt_directory(
        f"Destination folder (Enter for default: {mass_calc_dir}): ",
        mass_calc_dir,
    )
    default_csv_name = f"{source_csv.stem}_copy.csv"
    output_csv = _prompt_output_csv_path(output_dir, default_csv_name, disallow_path=source_csv)

    _out("")
    _out(f"Duplicating from: {source_csv}")
    _out(f"Output: {output_csv}")
    _out("")

    duplicate_csv(source_csv, output_csv)
    _out("\nCSV duplicated successfully")
    _out(f"Output saved to: {output_csv}")


if __name__ == "__main__":
    main()
