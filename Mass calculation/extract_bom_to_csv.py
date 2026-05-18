#!/usr/bin/env python3
"""
Extract component rows from BoM.xlsx into bom_extracted.csv.

Source sheet: BoM inverter card
Rows: data rows starting at row 3
"""

import csv
from pathlib import Path

import openpyxl


def extract_bom(excel_path, output_csv_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["BoM inverter card"]

    rows = []
    for r in range(3, ws.max_row + 1):
        designator = ws.cell(r, 4).value   # D
        if not designator:
            continue

        manufacturer = ws.cell(r, 5).value  # E
        part_number = ws.cell(r, 6).value   # F
        description = ws.cell(r, 7).value   # G
        quantity = ws.cell(r, 8).value      # H

        l_mm = ws.cell(r, 13).value         # M
        w_mm = ws.cell(r, 14).value         # N
        h_mm = ws.cell(r, 15).value         # O
        volume_cm3 = ws.cell(r, 16).value   # P
        density_min = ws.cell(r, 17).value  # Q
        density_max = ws.cell(r, 18).value  # R
        mass_existing = ws.cell(r, 21).value  # U

        rows.append([
            designator,
            manufacturer,
            part_number,
            description,
            quantity,
            l_mm,
            w_mm,
            h_mm,
            volume_cm3,
            density_min,
            density_max,
            mass_existing,
        ])

    with open(output_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Designators",
            "Manufacturer",
            "Part_Number",
            "Description",
            "Quantity",
            "L_mm",
            "W_mm",
            "H_mm",
            "Volume_cm3",
            "Density_min",
            "Density_max",
            "Mass_existing",
        ])
        writer.writerows(rows)

    return len(rows)


def main():
    excel_path = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing\BoM.xlsx")
    output_csv = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\bom_extracted.csv")

    count = extract_bom(excel_path, output_csv)
    print(f"Extracted components: {count}")
    print(f"Output file: {output_csv}")


if __name__ == "__main__":
    main()
