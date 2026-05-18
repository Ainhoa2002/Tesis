#!/usr/bin/env python3
"""
Synchronization Module for BOM Mass Calculations
Keeps original Excel and updated Excel in sync
"""

import openpyxl
import csv
from pathlib import Path
from datetime import datetime

class BOMSynchronizer:
    """Synchronize changes between Excel files and recalculate masses"""
    
    def __init__(self):
        self.original_excel = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing\BoM.xlsx")
        self.updated_excel = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\BoM_UPDATED.xlsx")
        self.extracted_csv = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\bom_extracted.csv")
        self.calculated_csv = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\bom_with_masses.csv")
    
    def extract_from_excel(self, excel_path, output_csv):
        """Extract BOM data from Excel to CSV"""
        print(f"Extracting data from: {excel_path}")
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        data_rows = []
        for row_idx in range(3, ws.max_row + 1):
            designator = ws.cell(row_idx, 4).value
            manufacturer = ws.cell(row_idx, 5).value
            part_number = ws.cell(row_idx, 6).value
            description = ws.cell(row_idx, 7).value
            quantity = ws.cell(row_idx, 8).value
            
            L = ws.cell(row_idx, 13).value
            W = ws.cell(row_idx, 14).value
            H = ws.cell(row_idx, 15).value
            volume = ws.cell(row_idx, 16).value
            density_min = ws.cell(row_idx, 17).value
            density_max = ws.cell(row_idx, 18).value
            mass = ws.cell(row_idx, 21).value
            
            if designator:
                data_rows.append([
                    designator, manufacturer, part_number, description, quantity,
                    L, W, H, volume, density_min, density_max, mass
                ])
        
        with open(output_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Designators', 'Manufacturer', 'Part_Number', 'Description', 'Quantity', 
                           'L_mm', 'W_mm', 'H_mm', 'Volume_cm3', 'Density_min', 'Density_max', 'Mass_existing'])
            writer.writerows(data_rows)
        
        print(f"✓ Extracted {len(data_rows)} components to {output_csv}")
        return len(data_rows)
    
    def compare_files(self):
        """Compare original and updated Excel files for differences"""
        print("\n" + "="*80)
        print("COMPARING FILES FOR CHANGES")
        print("="*80)
        
        wb_orig = openpyxl.load_workbook(self.original_excel, data_only=True)
        ws_orig = wb_orig.active
        
        wb_updated = openpyxl.load_workbook(self.updated_excel, data_only=True)
        ws_updated = wb_updated.active
        
        changes = []
        
        # Compare each row
        for row_idx in range(3, min(ws_orig.max_row, ws_updated.max_row) + 1):
            designator_orig = ws_orig.cell(row_idx, 4).value
            
            if designator_orig:
                # Check for changes in key columns
                for col in [13, 14, 15, 17, 18]:  # L, W, H, Density_min, Density_max
                    val_orig = ws_orig.cell(row_idx, col).value
                    val_updated = ws_updated.cell(row_idx, col).value
                    
                    if val_orig != val_updated:
                        col_names = {13: 'L', 14: 'W', 15: 'H', 17: 'Density_min', 18: 'Density_max'}
                        changes.append({
                            'designator': designator_orig,
                            'column': col_names.get(col, f'Col{col}'),
                            'original': val_orig,
                            'updated': val_updated
                        })
        
        if changes:
            print(f"\nFound {len(changes)} changes in updated file:\n")
            for change in changes:
                print(f"  {change['designator']} - {change['column']}: {change['original']} → {change['updated']}")
        else:
            print("\nNo changes detected between files.")
        
        print("=" * 80 + "\n")
        return changes
    
    def sync_to_original(self):
        """Copy calculated masses back to original Excel"""
        print("Syncing masses back to original Excel...")
        
        wb_updated = openpyxl.load_workbook(self.updated_excel, data_only=True)
        ws_updated = wb_updated.active
        
        wb_orig = openpyxl.load_workbook(self.original_excel)
        ws_orig = wb_orig.active
        
        updated_count = 0
        
        for row_idx in range(3, ws_updated.max_row + 1):
            designator = ws_updated.cell(row_idx, 4).value
            mass = ws_updated.cell(row_idx, 21).value
            
            if designator and mass and isinstance(mass, (int, float)):
                # Find matching row in original
                for orig_row_idx in range(3, ws_orig.max_row + 1):
                    orig_designator = ws_orig.cell(orig_row_idx, 4).value
                    if orig_designator == designator:
                        ws_orig.cell(orig_row_idx, 21).value = mass
                        updated_count += 1
                        break
        
        # Save with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"BoM_{timestamp}.xlsx"
        backup_path = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation") / backup_name
        wb_orig.save(str(backup_path))
        
        print(f"✓ Synced {updated_count} masses")
        print(f"✓ Backup saved: {backup_path}")
    
    def full_sync_workflow(self):
        """Complete synchronization workflow"""
        print("\n" + "#"*80)
        print("# FULL SYNCHRONIZATION WORKFLOW")
        print("#"*80 + "\n")
        
        # Step 1: Compare
        changes = self.compare_files()
        
        # Step 2: Extract from original if changes detected
        if changes:
            print("Re-extracting data from modified original Excel...")
            self.extract_from_excel(self.original_excel, self.extracted_csv)
        else:
            print("No changes detected. Using existing extracted data.")
        
        # Step 3: Re-calculate masses
        print("\nRe-calculating masses with updated data...")
        from bom_mass_calculator_excel import process_bom, write_output_csv
        results = process_bom(self.extracted_csv)
        write_output_csv(results, self.calculated_csv)
        
        print(f"✓ Processed {len(results)} components")
        
        # Step 4: Update the updated Excel
        print("\nUpdating BoM_UPDATED.xlsx with recalculated values...")
        self.load_and_write_masses(self.calculated_csv)
        
        print("\n" + "#"*80)
        print("# SYNCHRONIZATION COMPLETE")
        print("#"*80 + "\n")
        
        return results
    
    def load_and_write_masses(self, csv_path):
        """Load masses from CSV and write to updated Excel"""
        masses = {}
        
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                designator = row.get('Designators', '').strip()
                mass = row.get('Mass (g)', '').strip()
                if designator and mass:
                    try:
                        masses[designator] = float(mass)
                    except ValueError:
                        pass
        
        wb = openpyxl.load_workbook(self.updated_excel)
        ws = wb.active
        
        updated_count = 0
        for row_idx in range(3, ws.max_row + 1):
            designator_cell = ws.cell(row_idx, 4)
            designator = designator_cell.value
            
            if designator:
                des_key = str(designator).strip()
                mass_cell = ws.cell(row_idx, 21)
                
                if des_key in masses:
                    mass_val = masses[des_key]
                    mass_cell.value = round(mass_val, 4)
                    updated_count += 1
        
        wb.save(self.updated_excel)
        print(f"✓ Updated {updated_count} masses in BoM_UPDATED.xlsx")

def main():
    sync = BOMSynchronizer()
    
    print("\n" + "="*80)
    print("BOM SYNCHRONIZATION TOOL")
    print("="*80)
    
    print("\nOptions:")
    print("1. COMPARE files (check for differences)")
    print("2. FULL SYNC (extract, recalculate, update)")
    print("3. SYNC to original (copy back to source Excel)")
    
    choice = input("\nSelect option (1-3): ").strip()
    
    if choice == "1":
        sync.compare_files()
    elif choice == "2":
        results = sync.full_sync_workflow()
        print(f"\nTotal components: {len(results)}")
    elif choice == "3":
        sync.sync_to_original()
    else:
        print("Invalid option")

if __name__ == '__main__':
    main()
