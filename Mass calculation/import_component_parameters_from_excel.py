#!/usr/bin/env python3
"""
Import component parameters from Excel into CSV format.
Excel must have the same columns as the output CSV (Designators, Section, Subsection, etc.).
This allows building component databases directly in Excel and importing to the pipeline.
"""

import csv
from pathlib import Path
import openpyxl


def import_from_excel(workbook_path, output_csv, sheet_name="Sheet1"):
    """
    Import component parameters from Excel to CSV.
    
    Args:
        workbook_path: Path to Excel file with component data
        output_csv: Output CSV path
        sheet_name: Name of the sheet to read (default: Sheet1)
    
    Returns:
        Number of rows imported
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    
    try:
        ws = wb[sheet_name]
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in workbook")
        print(f"Available sheets: {wb.sheetnames}")
        return 0
    
    # Read header from first row
    header_row = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value is None:
            break
        header_row.append(str(cell_value).strip())
    
    if not header_row:
        print("Error: Excel file has no headers in first row")
        return 0
    
    print(f"Found {len(header_row)} columns: {', '.join(header_row[:5])}...")
    
    rows = []
    
    # Read data rows
    for r in range(2, ws.max_row + 1):
        row_data = {}
        has_content = False
        
        for col_idx, field_name in enumerate(header_row, start=1):
            cell_value = ws.cell(r, col_idx).value
            
            # Convert to string and strip whitespace
            if cell_value is not None:
                row_data[field_name] = str(cell_value).strip()
                has_content = True
            else:
                row_data[field_name] = ""
        
        # Skip completely empty rows
        if has_content:
            rows.append(row_data)
    
    # Write CSV with same headers as read from Excel
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=header_row)
        writer.writeheader()
        writer.writerows(rows)
    
    return len(rows)


def main():
    mass_calc_dir = Path(__file__).parent
    default_input_dir = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing")
    
    print("=" * 70)
    print("IMPORT COMPONENT PARAMETERS FROM EXCEL")
    print("=" * 70)
    print()
    
    # Ask for base directory where Excel workbook is located.
    while True:
        dir_prompt = f"📁 Folder path (Enter for default: {default_input_dir}): "
        folder_input = input(dir_prompt).strip()
        input_dir = Path(folder_input) if folder_input else default_input_dir

        if input_dir.exists() and input_dir.is_dir():
            break
        print(f"❌ Folder not found: {input_dir}")

    # Ask for workbook name (with or without .xlsx extension).
    while True:
        workbook_name = input("\n📘 Excel file name (e.g., BoM_fuse_card or BoM_fuse_card.xlsx): ").strip()
        if not workbook_name:
            print("❌ File name cannot be empty")
            continue

        if not workbook_name.lower().endswith(".xlsx"):
            workbook_name = workbook_name + ".xlsx"

        excel_path = input_dir / workbook_name
        if excel_path.exists():
            break
        print(f"❌ Excel file not found: {excel_path}")
    
    workbook_stem = excel_path.stem
    if workbook_stem.lower().startswith("bom_"):
        default_csv_name = f"{workbook_stem[4:]}_component_parameters.csv"
    else:
        default_csv_name = f"{workbook_stem}_component_parameters.csv"

    # Ask for output CSV file name.
    while True:
        csv_prompt = f"\n📄 Output CSV filename (Enter for default: {default_csv_name}): "
        csv_name = input(csv_prompt).strip()
        if not csv_name:
            csv_name = default_csv_name

        if not csv_name:
            print("❌ Filename cannot be empty")
            continue
        
        if not csv_name.endswith(".csv"):
            csv_name = csv_name + ".csv"
        
        output_csv = mass_calc_dir / csv_name
        break
    
    # Ask for sheet name
    sheet_name = input("\n📋 Sheet name (default: Sheet1): ").strip()
    if not sheet_name:
        sheet_name = "Sheet1"
    
    print()
    print(f"Importing from: {excel_path}")
    print(f"Sheet: {sheet_name}")
    print(f"Output: {output_csv}")
    print()
    
    count = import_from_excel(excel_path, output_csv, sheet_name)
    
    print(f"\n✅ Imported {count} component rows")
    print(f"✅ Output saved to: {output_csv}")


if __name__ == "__main__":
    main()
