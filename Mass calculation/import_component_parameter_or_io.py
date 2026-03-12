#!/usr/bin/env python3
"""
Import component parameters from Excel into CSV format or it copies changing the name another .csv file (useful for input and output flows to can edit them).
Excel must have the same columns as the output CSV (Designators, Section, Subsection, etc.).
This allows building component databases directly in Excel and importing to the pipeline.
"""

import csv
import shutil
from pathlib import Path
import openpyxl


def _is_yes_answer(answer):
    return answer.strip().casefold() in {"y", "yes", "s", "si", "sí"}


def _prompt_import_mode():
    print("\nImport options:")
    print("1. Excel (BoM parameters)")
    print("2. Duplicate .csv")

    while True:
        option = input("Choose option [1/2]: ").strip().lower()
        if option in {"1", "excel", "x"}:
            return "excel"
        if option in {"2", "csv", "duplicate", "dup"}:
            return "csv"
        print("Invalid option. Enter 1 or 2.")


def _prompt_directory(prompt_text, default_dir):
    while True:
        dir_input = input(prompt_text).strip()
        chosen_dir = Path(dir_input) if dir_input else default_dir
        if chosen_dir.exists() and chosen_dir.is_dir():
            return chosen_dir
        print(f"Folder not found: {chosen_dir}")


def _prompt_output_csv_path(default_dir, default_name, disallow_path=None):
    while True:
        csv_prompt = f"\nOutput CSV filename (Enter for default: {default_name}): "
        csv_name = input(csv_prompt).strip()
        if not csv_name:
            csv_name = default_name

        if not csv_name:
            print("Filename cannot be empty")
            continue

        if not csv_name.endswith(".csv"):
            csv_name = csv_name + ".csv"

        output_csv = default_dir / csv_name
        if disallow_path is not None and output_csv.resolve() == disallow_path.resolve():
            print("Output CSV must be different from source CSV.")
            continue

        if output_csv.exists():
            overwrite_answer = input(
                f"File already exists: {output_csv}. Overwrite? [y/N]: "
            )
            if not _is_yes_answer(overwrite_answer):
                print("Choose a different filename.")
                continue
        return output_csv


def _resolve_source_csv(source_dir, source_name, extra_search_dirs=None):
    if not source_name.lower().endswith(".csv"):
        source_name = source_name + ".csv"

    source_path = Path(source_name)
    if source_path.is_absolute():
        return source_path if source_path.exists() and source_path.is_file() else None

    primary_candidate = source_dir / source_path
    if primary_candidate.exists() and primary_candidate.is_file():
        return primary_candidate

    for search_dir in extra_search_dirs or []:
        fallback_candidate = search_dir / source_path
        if fallback_candidate.exists() and fallback_candidate.is_file():
            return fallback_candidate

    return None


def _prompt_source_csv_path(default_dir, extra_search_dirs=None):
    while True:
        dir_prompt = f"Source CSV folder (Enter for default: {default_dir}): "
        source_dir = _prompt_directory(dir_prompt, default_dir)

        source_name = input("Source CSV filename: ").strip()
        if not source_name:
            print("Filename cannot be empty")
            continue

        source_csv = _resolve_source_csv(source_dir, source_name, extra_search_dirs)
        if source_csv is not None:
            if source_csv.parent.resolve() != source_dir.resolve():
                print(f"Source CSV found in: {source_csv.parent}")
            return source_csv

        filename = source_name if source_name.lower().endswith(".csv") else f"{source_name}.csv"
        print(f"CSV file not found in: {source_dir / filename}")
        if extra_search_dirs:
            print("Also checked fallback folders.")


def duplicate_csv(source_csv, output_csv):
    shutil.copyfile(source_csv, output_csv)


def _select_sheet_name(workbook, requested_name=None):
    if requested_name:
        if requested_name in workbook.sheetnames:
            return requested_name
        raise KeyError(requested_name)

    if "Parameters" in workbook.sheetnames:
        return "Parameters"
    if "Sheet1" in workbook.sheetnames:
        return "Sheet1"
    if len(workbook.sheetnames) == 1:
        return workbook.sheetnames[0]
    return workbook.sheetnames[0]


def import_from_excel(workbook_path, output_csv, sheet_name=None):
    """
    Import component parameters from Excel to CSV.
    -Path to Excel file with component data
    -Output CSV path
    -sheet_name: Name of the sheet to read (default: auto-detect)
    
    Returns:
        Tuple with imported row count and the sheet name used.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    
    try:
        resolved_sheet_name = _select_sheet_name(wb, sheet_name)
        ws = wb[resolved_sheet_name]
    except KeyError:
        requested_label = sheet_name if sheet_name else "<auto-detect>"
        print(f"Error: Sheet '{requested_label}' could not be resolved in workbook")
        print(f"Available sheets: {wb.sheetnames}")
        return 0, None
    
    # Read header from first row
    header_row = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value is None:
            break
        header_row.append(str(cell_value).strip())
    
    if not header_row:
        print("Error: Excel file has no headers in first row")
        return 0, None
    
    print(f"Found {len(header_row)} columns: {', '.join(header_row[:5])}...")
    
    rows = []
    
    # Read data rows
    for r in range(2, ws.max_row + 1):
        row_data = {}
        has_content = False
        
        for col_idx, field_name in enumerate(header_row, start=1):
            cell_value = ws.cell(r, col_idx).value
            
            # Convert to string and strip whitespace
            if cell_value is not None:
                row_data[field_name] = str(cell_value).strip()
                has_content = True
            else:
                row_data[field_name] = ""
        
        # Skip completely empty rows
        if has_content:
            rows.append(row_data)
    
    # Write CSV with same headers as read from Excel
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=header_row)
        writer.writeheader()
        writer.writerows(rows)
    
    return len(rows), resolved_sheet_name


def main():
    mass_calc_dir = Path(__file__).parent
    current_git_dir = Path.cwd().resolve()
    default_input_dir = current_git_dir
    
    print("IMPORT COMPONENT PARAMETERS")
    mode = _prompt_import_mode()

    if mode == "excel":
        # Ask for base directory where Excel workbook is located.
        input_dir = _prompt_directory(
            f"Folder path (Enter for default: {default_input_dir}): ",
            default_input_dir,
        )

        # Ask for workbook name (with or without .xlsx extension).
        while True:
            workbook_name = input("\n Excel file name (e.g., BoM_fuse_card or BoM_fuse_card.xlsx): ").strip()
            if not workbook_name:
                print("File name cannot be empty")
                continue

            if not workbook_name.lower().endswith(".xlsx"):
                workbook_name = workbook_name + ".xlsx"

            excel_path = input_dir / workbook_name
            if excel_path.exists():
                break
            print(f"Excel file not found: {excel_path}")

        workbook_stem = excel_path.stem
        if workbook_stem.lower().startswith("bom_"):
            default_csv_name = f"{workbook_stem[4:]}_component_parameters.csv"
        else:
            default_csv_name = f"{workbook_stem}_component_parameters.csv"

        output_csv = _prompt_output_csv_path(current_git_dir, default_csv_name)

        # Ask for sheet name
        sheet_name = input("\nSheet name (Enter for auto-detect): ").strip()
        if not sheet_name:
            sheet_name = None

        print()
        print(f"Importing from: {excel_path}")
        print(f"Sheet: {sheet_name or 'auto-detect'}")
        print(f"Output: {output_csv}")
        print()

        count, resolved_sheet_name = import_from_excel(excel_path, output_csv, sheet_name)

        if resolved_sheet_name is None:
            print("\nNo CSV file was generated.")
            return

        print(f"\nImported {count} component rows")
        print(f"Sheet used: {resolved_sheet_name}")
        print(f"Output saved to: {output_csv}")
        return

    source_csv = _prompt_source_csv_path(
        current_git_dir,
        extra_search_dirs=[mass_calc_dir],
    )
    output_dir = _prompt_directory(
        f"Destination folder (Enter for default: {current_git_dir}): ",
        current_git_dir,
    )
    default_csv_name = f"{source_csv.stem}_copy.csv"
    output_csv = _prompt_output_csv_path(output_dir, default_csv_name, disallow_path=source_csv)

    print()
    print(f"Duplicating from: {source_csv}")
    print(f"Output: {output_csv}")
    print()

    duplicate_csv(source_csv, output_csv)
    print("\nCSV duplicated successfully")
    print(f"Output saved to: {output_csv}")


if __name__ == "__main__":
    main()
