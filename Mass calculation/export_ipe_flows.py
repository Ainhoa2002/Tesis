#!/usr/bin/env python3
"""
Export IPE flows for ecoinvent from the workbook BOM data.

This script reads:
- Sheet: BoM inverter card
- Flow name from column Y (Name)
- Unit from column I (Unit)
- Total quantity from column K (Total quantity)

If K has invalid or missing value, it recomputes row total as H * J,
where J is recomputed from P,Q,R,S when possible.

Output CSV columns:
Flow,Unit,Amount,Direction
"""

import csv
from collections import OrderedDict
from pathlib import Path

import openpyxl


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", ".")
    if text == "":
        return None

    # Parse simple ranges like 0.35-0.4 by averaging.
    if "-" in text and not text.startswith("-"):
        parts = text.split("-")
        if len(parts) == 2:
            try:
                return (float(parts[0]) + float(parts[1])) / 2.0
            except ValueError:
                return None

    try:
        return float(text)
    except ValueError:
        return None


def recompute_j_kg_per_element(row_values):
    # Excel columns: L=12, P=16, Q=17, R=18, S=19 in 1-based indexing
    # row_values here is a 0-based list for columns A..Y.
    l_flag = row_values[11]
    if str(l_flag).strip().upper() != "NO":
        return None

    p = to_float(row_values[15])
    q = to_float(row_values[16])
    r = to_float(row_values[17])
    s = to_float(row_values[18])

    if p is None or q is None or r is None:
        return None

    s = 0.0 if s is None else s
    return (p * ((q + r) / 2.0) + s) / 1000.0


def recompute_total_kg(row_values, j_cached, k_cached):
    # Prefer cached K from workbook if valid.
    k = to_float(k_cached)
    if k is not None:
        return k

    h = to_float(row_values[7])
    if h is None:
        return None

    j = to_float(j_cached)

    # Special fixed value present in workbook for TRV line (IC3-IC6).
    if j is None and str(row_values[3]).strip().startswith("IC3"):
        j = 4.8 / 1000.0

    if j is None:
        j = recompute_j_kg_per_element(row_values)

    if j is None:
        return None

    return h * j


def export_ipe_flows(workbook_path, output_csv_path):
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)

    ws_formulas = wb_formulas["BoM inverter card"]
    ws_values = wb_values["BoM inverter card"]

    grouped = OrderedDict()

    for r in range(3, ws_formulas.max_row + 1):
        flow_name = ws_values.cell(r, 25).value  # Y
        unit = ws_values.cell(r, 9).value        # I

        if not flow_name or not unit:
            continue

        row_values = [ws_values.cell(r, c).value for c in range(1, 26)]
        j_cached = ws_values.cell(r, 10).value
        k_cached = ws_values.cell(r, 11).value

        total_kg = recompute_total_kg(row_values, j_cached, k_cached)
        if total_kg is None:
            continue

        key = (str(flow_name).strip(), str(unit).strip())
        grouped[key] = grouped.get(key, 0.0) + total_kg

    with open(output_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Flow", "Unit", "Amount", "Direction"])
        for (flow, unit), amount in grouped.items():
            writer.writerow([flow, unit, round(amount, 12), "Input"])

    return grouped


def main():
    workbook = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing\BoM.xlsx")
    output_csv = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\ipe_flows_from_bom.csv")

    grouped = export_ipe_flows(workbook, output_csv)

    print(f"Exported flows: {len(grouped)}")
    print(f"Total amount sum: {sum(grouped.values()):.12f}")
    print(f"Output file: {output_csv}")


if __name__ == "__main__":
    main()
