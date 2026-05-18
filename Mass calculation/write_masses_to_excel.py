#!/usr/bin/env python3
"""
Write calculated masses back to Excel BOM
Maps calculated data back to original Excel structure
"""

import openpyxl
import csv
from pathlib import Path


def normalize_designator(text):
    if text is None:
        return ""
    normalized = str(text).strip().upper()
    normalized = normalized.replace("–", "-").replace("—", "-")
    normalized = " ".join(normalized.split())
    return normalized

def load_calculated_masses(csv_path):
    """Load calculated masses from CSV"""
    masses = {}
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            designator = row.get('Designators', '').strip()
            mass = row.get('Mass (g)', '').strip()
            if designator and mass:
                try:
                    masses[normalize_designator(designator)] = float(mass)
                except ValueError:
                    pass
    
    return masses

def update_excel_with_masses(excel_path, calculated_masses, output_path):
    """Update Excel file with calculated masses"""
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    updated_count = 0
    
    # Process each data row (starts from row 3)
    for row_idx in range(3, ws.max_row + 1):
        # Col D = Designators
        designator_cell = ws.cell(row_idx, 4)
        designator = designator_cell.value
        
        if designator:
            # Normalize designator for lookup
            des_key = normalize_designator(designator)
            
            # Col U = Mass (g)
            mass_cell = ws.cell(row_idx, 21)
            
            # Check if we have a calculated mass for this component
            if des_key in calculated_masses:
                mass_val = calculated_masses[des_key]
                # Format to 4 decimal places
                mass_cell.value = round(mass_val, 4)
                updated_count += 1
                print(f"Updated {des_key}: {mass_val:.4f} g")
    
    # Save the updated workbook
    wb.save(output_path)
    print(f"\nUpdated {updated_count} rows")
    print(f"Saved to: {output_path}")

def main():
    original_excel = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing\BoM.xlsx")
    calculated_csv = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\bom_with_masses.csv")
    output_excel = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\BoM_UPDATED.xlsx")
    
    print("Loading calculated masses...")
    masses = load_calculated_masses(calculated_csv)
    print(f"Loaded {len(masses)} calculated masses\n")
    
    print("Updating Excel file...")
    update_excel_with_masses(original_excel, masses, output_excel)

if __name__ == '__main__':
    main()
