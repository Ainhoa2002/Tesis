#!/usr/bin/env python3
"""
Bootstrap inverter_power_card_component_parameters.csv from BoM.xlsx.
Use this once, then edit inverter_power_card_component_parameters.csv directly.
"""

import csv
from pathlib import Path

import openpyxl


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def seed_from_excel(workbook_path, output_csv):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["BoM inverter card"]

    rows = []

    for r in range(3, ws.max_row + 1):
        designators = ws.cell(r, 4).value  # D
        if not designators:
            continue

        manufacturer = ws.cell(r, 5).value
        part_number = ws.cell(r, 6).value
        description = ws.cell(r, 7).value

        number_elements = ws.cell(r, 8).value   # H
        unit = ws.cell(r, 9).value              # I
        qty_per_element = ws.cell(r, 10).value  # J

        datasheet_flag = ws.cell(r, 12).value   # L
        has_datasheet_info = "NO" if _clean_text(datasheet_flag).upper() == "NO" else "YES"

        l_mm = ws.cell(r, 13).value             # M
        w_mm = ws.cell(r, 14).value             # N
        h_mm = ws.cell(r, 15).value             # O
        volume_cm3_excel = ws.cell(r, 16).value # P

        metal_extra_g = ws.cell(r, 19).value    # S
        other_extra_g = ws.cell(r, 20).value    # T

        # In this workbook, U contains the selected similar component title.
        component_database_title = ws.cell(r, 21).value  # U
        _mass_datasheet = ws.cell(r, 22).value           # V

        _scale_with_mass = ws.cell(r, 23).value          # W
        database = ws.cell(r, 24).value                  # X
        ecoinvent_flow = ws.cell(r, 25).value            # Y
        _other_possible_models = ws.cell(r, 26).value    # Z
        comments = ws.cell(r, 27).value                  # AA
        _reliability = ws.cell(r, 28).value              # AB
        _completeness = ws.cell(r, 29).value             # AC
        _temporal_correlation = ws.cell(r, 30).value     # AD

        rows.append(
            {
                "Designators": designators,
                "Manufacturer": manufacturer,
                "Part_Number": part_number,
                "Description": description,
                "number_elements": number_elements,
                "unit": unit,
                "Quantity_per_element": qty_per_element,
                "Has_datasheet_info": has_datasheet_info,
                "L_mm": l_mm,
                "W_mm": w_mm,
                "H_mm": h_mm,
                "Volume_cm3_excel": volume_cm3_excel,
                "Density_min_g_cm3": ws.cell(r, 17).value,
                "Density_max_g_cm3": ws.cell(r, 18).value,
                "Metal_extra_g": metal_extra_g,
                "Other_extra_g": other_extra_g,
                "Database": database,
                "Ecoinvent_flow": ecoinvent_flow,
                "Ecoinvent_unit": unit,
                "Direction": "Input",
                "Ecoinvent_amount_override": "",
                "Database_component_title": component_database_title,
                "Comments": comments,
                "Notes": "",
            }
        )

    fieldnames = [
        "Designators",
        "Manufacturer",
        "Part_Number",
        "Description",
        "number_elements",
        "unit",
        "Quantity_per_element",
        "Has_datasheet_info",
        "L_mm",
        "W_mm",
        "H_mm",
        "Volume_cm3_excel",
        "Density_min_g_cm3",
        "Density_max_g_cm3",
        "Metal_extra_g",
        "Other_extra_g",
        "Database",
        "Ecoinvent_flow",
        "Ecoinvent_unit",
        "Direction",
        "Ecoinvent_amount_override",
        "Database_component_title",
        "Comments",
        "Notes",
    ]

    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


def main():
    workbook = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing\BoM.xlsx")
    output_csv = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\inverter_power_card_component_parameters.csv")

    count = seed_from_excel(workbook, output_csv)
    print(f"Seeded parameter rows: {count}")
    print(f"Output file: {output_csv}")


if __name__ == "__main__":
    main()
