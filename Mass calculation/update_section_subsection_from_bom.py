#!/usr/bin/env python3
"""
Update Section and Subsection in inverter_power_card_component_parameters.csv
from the BoM.xlsx file.
"""

import csv
from pathlib import Path
import openpyxl


def load_bom_mapping(workbook_path):
    """Load Section and Subsection from BoM by Designators."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["BoM inverter card"]
    
    mapping = {}  # designators -> (section, subsection)
    
    for r in range(3, ws.max_row + 1):
        category = ws.cell(r, 1).value            # A
        section = ws.cell(r, 2).value             # B
        subsection = ws.cell(r, 3).value          # C
        designators = ws.cell(r, 4).value         # D
        
        if not designators:
            continue
        
        # Clean values
        section = str(section).strip() if section else ""
        subsection = str(subsection).strip() if subsection else ""
        designators = str(designators).strip()
        
        mapping[designators] = (section, subsection)
    
    return mapping


def update_csv_from_bom(csv_path, bom_mapping):
    """Update CSV Section and Subsection columns from BOM mapping."""
    rows = []
    updated_count = 0
    missing_designators = []
    
    # Read current CSV
    with open(csv_path, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        
        for row in reader:
            designators = row.get('Designators', '').strip()
            
            if designators in bom_mapping:
                section, subsection = bom_mapping[designators]
                row['Section'] = section
                row['Subsection'] = subsection
                updated_count += 1
            elif designators:
                missing_designators.append(designators)
            
            rows.append(row)
    
    # Write back CSV
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    
    return updated_count, missing_designators


def main():
    bom_path = Path(r"C:\Users\alorzaga\cernbox\WINDOWS\Desktop\TESIS\Power converters\Power converter\Manufacturing\BoM.xlsx")
    csv_path = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation\inverter_power_card_component_parameters.csv")
    
    if not bom_path.exists():
        print(f"Error: BoM file not found: {bom_path}")
        return
    
    if not csv_path.exists():
        print(f"Error: CSV file not found: {csv_path}")
        return
    
    print("Loading BoM mapping...")
    mapping = load_bom_mapping(bom_path)
    print(f"Loaded {len(mapping)} entries from BoM")
    
    print(f"\nUpdating {csv_path}...")
    updated_count, missing = update_csv_from_bom(csv_path, mapping)
    
    print(f"\n✅ Updated {updated_count} rows with Section and Subsection")
    
    if missing:
        print(f"\n⚠️  {len(missing)} designators not found in BoM:")
        for des in missing[:10]:
            print(f"   - {des}")
        if len(missing) > 10:
            print(f"   ... and {len(missing) - 10} more")


if __name__ == "__main__":
    main()
