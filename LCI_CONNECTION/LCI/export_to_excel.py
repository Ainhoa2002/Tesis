#!/usr/bin/env python3
"""Export subsystem results to Excel workbooks.

This script reuses subsystem selection from add_eliminate_component.py.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from add_eliminate_component import (
    MAX_SELECTION_ATTEMPTS,
    SelectionAborted,
    choose_subsystem,
    discover_subsystem_files,
)

BASE_DIR = Path(r"c:\Users\alorzaga\Git\tesis\Mass calculation")
DEFAULT_EXPORT_DIR = BASE_DIR.parent


def load_csv_optional(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    if not path.exists():
        return [], []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    return headers, rows


def write_sheet(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def prompt_output_directory(default_dir: Path) -> Path:
    while True:
        folder_input = input(
            f"\nOutput folder (Enter for default: {default_dir}): "
        ).strip()

        if not folder_input:
            return default_dir

        chosen_dir = Path(folder_input).expanduser()
        if not chosen_dir.is_absolute():
            chosen_dir = (Path.cwd() / chosen_dir).resolve()

        if chosen_dir.exists() and chosen_dir.is_dir():
            return chosen_dir

        print(f"Folder not found: {chosen_dir}")


def prompt_output_filename(default_name: str) -> str:
    while True:
        file_input = input(
            f"Output filename (Enter for default: {default_name}): "
        ).strip()

        if not file_input:
            return default_name

        if not file_input.lower().endswith(".xlsx"):
            file_input = file_input + ".xlsx"

        return file_input


def choose_export_mode() -> str:
    print("\nExport mode:")
    print("  1. Export one subsystem")
    print("  2. Export all subsystems")

    attempts = 0
    while True:
        raw = input("Mode [1/2]: ").strip().lower()
        if raw in {"1", "one", "single", "subsystem"}:
            return "one"
        if raw in {"2", "all", "todo", "todos", "*"}:
            return "all"

        attempts += 1
        print("Invalid option. Enter 1 or 2.")
        if attempts >= MAX_SELECTION_ATTEMPTS:
            raise SelectionAborted("Too many invalid attempts. Operation canceled.")


def export_all_subsystems_to_excel(
    base_dir: Path,
    subsystems: Dict[str, Path],
    output_dir: Path,
) -> Tuple[List[Path], List[str]]:
    exported_paths: List[Path] = []
    skipped_subsystems: List[str] = []
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("\nChoose output filename for each subsystem (Enter for default).")
    for subsystem_name in sorted(subsystems.keys()):
        default_filename = f"{subsystem_name}_results_export_{stamp}.xlsx"
        print(f"\nSubsystem: {subsystem_name}")
        output_filename = prompt_output_filename(default_filename)

        output = export_subsystem_results_to_excel(
            base_dir,
            subsystem_name,
            output_dir,
            output_filename,
        )

        if output is None:
            skipped_subsystems.append(subsystem_name)
            continue

        exported_paths.append(output)

    return exported_paths, skipped_subsystems


def export_subsystem_results_to_excel(
    base_dir: Path,
    subsystem: str,
    output_dir: Path,
    output_filename: str,
) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl is required for Excel export.")
        print("Install it with: .\\.venv\\Scripts\\python.exe -m pip install openpyxl")
        return None

    sources = [
        ("Parameters", base_dir / f"{subsystem}_component_parameters.csv"),
        ("Mass_Results", base_dir / f"{subsystem}_component_mass_results.csv"),
        ("Component_IO", base_dir / f"{subsystem}_component_io_flows.csv"),
        ("Grouped_Flows", base_dir / f"{subsystem}_ipe_flows_from_parameters.csv"),
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    exported = 0

    for sheet_name, csv_path in sources:
        headers, rows = load_csv_optional(csv_path)
        if not headers:
            continue
        ws = workbook.create_sheet(sheet_name[:31])
        write_sheet(ws, headers, rows)
        exported += 1

    if exported == 0:
        print("No data found to export for this subsystem.")
        return None

    output_path = output_dir / output_filename
    workbook.save(output_path)
    return output_path


def export_total_bom_to_excel(
    base_dir: Path,
    output_dir: Path,
    output_filename: str,
) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl is required for Excel export.")
        print("Install it with: .\\.venv\\Scripts\\python.exe -m pip install openpyxl")
        return None

    sources = [
        ("Parameters_All", base_dir / "component_library_parameters_all.csv"),
        ("Mass_Results_All", base_dir / "component_library_mass_results_all.csv"),
        ("Ecoinvent_Totals", base_dir / "component_library_ecoinvent_totals.csv"),
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    exported = 0

    for sheet_name, csv_path in sources:
        headers, rows = load_csv_optional(csv_path)
        if not headers:
            continue
        ws = workbook.create_sheet(sheet_name[:31])
        write_sheet(ws, headers, rows)
        exported += 1

    if exported == 0:
        print("No data found to export for total BoM.")
        return None

    output_path = output_dir / output_filename
    workbook.save(output_path)
    return output_path


def write_export_readme(
    output_dir: Path,
    exported_items: List[Tuple[Path, str]],
) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    readme_path = output_dir / f"export_readme_{stamp}.txt"
    lines = [
        "# Resumen de export",
        "",
        "Archivos exportados:",
    ]

    for file_path, description in exported_items:
        lines.append(f"- {file_path.name}: {description}")

    readme_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return readme_path


def main() -> None:
    try:
        subsystems = discover_subsystem_files(BASE_DIR)
        export_mode = choose_export_mode()

        if export_mode == "one":
            subsystem_name, _ = choose_subsystem(subsystems)
            export_dir = prompt_output_directory(DEFAULT_EXPORT_DIR)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"{subsystem_name}_results_export_{stamp}.xlsx"
            export_filename = prompt_output_filename(default_filename)

            output = export_subsystem_results_to_excel(
                BASE_DIR,
                subsystem_name,
                export_dir,
                export_filename,
            )
            if output is not None:
                print(f"\nExport completed: {output}")
                readme_path = write_export_readme(
                    export_dir,
                    [
                        (
                            output,
                            "Excel de subsistema con hojas: Parameters, Mass_Results, Component_IO, Grouped_Flows.",
                        )
                    ],
                )
                print(f"Readme creado: {readme_path}")
            return

        export_dir = prompt_output_directory(DEFAULT_EXPORT_DIR)
        exported_paths, skipped_subsystems = export_all_subsystems_to_excel(
            BASE_DIR,
            subsystems,
            export_dir,
        )

        print("\nBoM total export file (includes parameters_all, mass_results_all and ecoinvent_totals):")
        bom_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_bom_filename = f"bom_total_export_{bom_stamp}.xlsx"
        bom_filename = prompt_output_filename(default_bom_filename)
        bom_output = export_total_bom_to_excel(
            BASE_DIR,
            export_dir,
            bom_filename,
        )

        if exported_paths:
            print("\nExport completed for:")
            for path in exported_paths:
                print(f"  - {path}")
        else:
            print("\nNo subsystem exports were generated.")

        if bom_output is not None:
            print("\nBoM total export completed:")
            print(f"  - {bom_output}")
        else:
            print("\nBoM total export was not generated.")

        if skipped_subsystems:
            print("\nSubsystems with no exportable data:")
            for subsystem_name in skipped_subsystems:
                print(f"  - {subsystem_name}")

        readme_items: List[Tuple[Path, str]] = []
        for subsystem_output in exported_paths:
            readme_items.append(
                (
                    subsystem_output,
                    "Excel de subsistema con hojas: Parameters, Mass_Results, Component_IO, Grouped_Flows.",
                )
            )

        if bom_output is not None:
            readme_items.append(
                (
                    bom_output,
                    "Excel de BoM total con hojas: Parameters_All, Mass_Results_All y Ecoinvent_Totals.",
                )
            )

        if readme_items:
            readme_path = write_export_readme(export_dir, readme_items)
            print(f"\nReadme creado: {readme_path}")
    except SelectionAborted as exc:
        print(str(exc))


if __name__ == "__main__":
    main()
